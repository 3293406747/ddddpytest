from openpyxl import load_workbook
from pathlib import Path

path = Path(__file__).resolve().parent.parent.parent

def read_excel(filename,sheet=None):
	""" 读取excel测试数据 """
	filename = path.joinpath("data").joinpath(filename)
	wb = load_workbook(filename=filename,read_only=True)
	ws = wb[sheet] if sheet else wb.active
	column = next(ws.iter_rows(values_only=True))
	dataset = [dict(zip(column, row)) for row in ws.iter_rows(min_row=2,values_only=True)]
	return dataset