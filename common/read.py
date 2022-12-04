import json
from string import Template
import yaml,csv
from pathlib import Path
from openpyxl import load_workbook
from common.case import verifyCase, renderTemplate


path = Path(__file__).resolve().parent.parent
instance = {}

def read_config(filename="local.yaml", encoding="utf-8") ->dict:
	""" 读取配置文件 """
	if not instance.get(filename):
		with path.joinpath("config").joinpath(filename).open(encoding=encoding) as f:
			data = yaml.load(stream=f, Loader=yaml.FullLoader)
		instance[filename] = data
	config = renderTemplate(instance[filename])
	return config

def read_csv(filename, encoding='utf-8'):
	""" 读取csv测试数据 已废弃 """
	with path.joinpath("data").joinpath(filename).open(encoding=encoding) as f:
		reader = csv.reader(f)
		column = next(reader)
		dataset = [dict(zip(column,row)) for row in reader]
	return dataset

def read_excel(filename,sheet=None):
	""" 读取excel测试数据 """
	filename = path.joinpath("data").joinpath(filename)
	wb = load_workbook(filename=filename,read_only=True)
	ws = wb[sheet] if sheet else wb.active
	column = next(ws.iter_rows(values_only=True))
	dataset = [dict(zip(column, row)) for row in ws.iter_rows(min_row=2,values_only=True)]
	return dataset

def read_yaml(filename, encoding='utf-8'):
	""" 读取测试用例 """
	with path.joinpath("testcase").joinpath(filename).open(encoding=encoding) as f:
		case = yaml.load(stream=f, Loader=yaml.FullLoader)
	return case

def read_testcase(filename, item=0, encoding="utf-8"):
	""" 读取测试用例 """
	caseinfo = read_yaml(filename=filename, encoding=encoding)[item]
	caseinfo = verifyCase(caseinfo)
	if not caseinfo.get("data_path"):
		return [caseinfo]
	data_path = caseinfo.pop("data_path")
	sheet = caseinfo.pop("data_sheet") if caseinfo.get("data_sheet") else None
	caseinfo = json.dumps(caseinfo, ensure_ascii=False)
	data = read_excel(filename=data_path,sheet=sheet)
	caseList = [yaml.load(stream=Template(caseinfo).safe_substitute(i), Loader=yaml.FullLoader) for i in data]
	return caseList
