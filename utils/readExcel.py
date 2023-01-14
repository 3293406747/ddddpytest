from openpyxl import load_workbook


def readExcel(file, sheet=None):
	""" 读取excel文件 """
	wb = load_workbook(filename=file, read_only=True)
	ws = wb[sheet] if sheet else wb.active
	column = next(ws.iter_rows(values_only=True))
	return [dict(zip(column, row)) for row in ws.iter_rows(min_row=2, values_only=True)]
